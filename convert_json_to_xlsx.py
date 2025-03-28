"""转换 json 为 xlsx"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet
from scraper_utils.constants.workbook_style import (
    HYPERLINK_FONT,
    TEXT_CENTER_ALIGNMENT,
    RED_BOLD_FONT,
    YELLOW_FILL,
)
from scraper_utils.utils.json_util import read_json_sync
from scraper_utils.utils.workbook_util import write_workbook_sync, column_int2str as i2s

cwd = Path.cwd()


def main():
    ########## 输入 ##########
    date = '0328'
    category = 'dispensere'
    # category = 'veioze-si-lampi'
    save_path = cwd / f'temp/{category}-{date}.xlsx'
    ########## 输入 ##########

    wb, ws = create_workbook_template()
    products = read_product_json(Path(f'temp/{category}/{date}'))
    save_to_xlsx(wb, ws, products, save_path)


def create_workbook_template():
    """创建表格模板"""
    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore

    # 标题行
    ws['A1'] = 'pnk'
    ws['B1'] = '详情页链接'
    ws['C1'] = '标题'
    ws['D1'] = '类目'
    ws['E1'] = '来源链接'
    ws['F1'] = '排名'
    ws['G1'] = '产品图'
    ws['H1'] = '价格'
    ws['I1'] = 'Top 标志'
    ws['J1'] = '评分'
    ws['K1'] = '评论数'
    ws['L1'] = '最大可加购数'

    # 设置标题行样式
    for c in range(1, 12 + 1):
        ws.cell(1, c).fill = YELLOW_FILL
        ws.cell(1, c).font = RED_BOLD_FONT
        ws.cell(1, c).alignment = TEXT_CENTER_ALIGNMENT
        ws.column_dimensions[i2s(c)].width = int(120 / 7)

    return wb, ws


def read_product_json(json_dir: Path):
    """读取产品的 json 数据"""
    result: list[dict[str, None | str | int | float | bool]] = list()

    files = json_dir.glob('*.json')

    for f in files:
        # 读取 json 文件
        data: list[dict[str, None | str | int | float | bool]] = read_json_sync(f)

        # 筛选加购失败的记录、保留需要的字段、将清理结果添加到 result
        for d in data:
            # 筛选加购失败的记录
            if d['cart_added'] is not True:
                continue

            # 保留需要的字段
            result.append(
                {
                    'pnk': d['pnk'],
                    'title': d['title'],
                    'category': d['category'],
                    'rank': d['rank_in_category'],
                    'source_url': d['source_url'],
                    'detail_url': d['detail_url'],
                    'image_url': d['image_url'],
                    'price': d['price'],
                    'top_favorite': d['top_favorite'],
                    'rating': d['rating'],
                    'review': d['review'],
                    'max_qty': d['max_qty'],
                }
            )

    # 按 rank 排序
    result.sort(key=lambda _: _['rank'])  # type: ignore

    return result.copy()


def save_to_xlsx(
    wb: Workbook,
    ws: Worksheet,
    data: list[dict[str, None | str | int | float | bool]],
    save_path: str | Path,
):
    """写入数据到表格，并保存表格为文件"""

    for row, p in enumerate(data, 2):
        # pnk
        pnk: str = p['pnk']  # type: ignore
        ws[f'A{row}'] = pnk

        # 详情页链接
        # detail_url = build_product_url(pnk)
        detail_url: str = p['detail_url']  # type: ignore
        ws[f'B{row}'] = detail_url
        ws[f'B{row}'].hyperlink = Hyperlink(ref=detail_url, target=detail_url)
        ws[f'B{row}'].font = HYPERLINK_FONT

        # 标题
        title: str = p['title']  # type: ignore
        ws[f'C{row}'] = title

        # 类目
        category: str = p['category']  # type: ignore
        ws[f'D{row}'] = category

        # 来源链接
        source_url: str = p['source_url']  # type: ignore
        ws[f'E{row}'] = source_url
        ws[f'E{row}'].hyperlink = Hyperlink(ref=source_url, target=source_url)
        ws[f'E{row}'].font = HYPERLINK_FONT

        # 排名
        rank: int = p['rank']  # type: ignore
        ws[f'F{row}'] = rank

        # 产品图
        image_url: str | None = p['image_url']  # type: ignore
        if image_url is not None:
            ws[f'G{row}'] = image_url
            ws[f'G{row}'].hyperlink = Hyperlink(ref=image_url, target=image_url)
            ws[f'G{row}'].font = HYPERLINK_FONT
        else:
            ws[f'G{row}'] = '/'

        # 价格
        price: float = p['price']  # type: ignore
        ws[f'H{row}'] = price

        # Top 标志
        top_favorite: bool = p['top_favorite']  # type: ignore
        ws[f'I{row}'] = '是' if top_favorite else '否'

        # 评分
        rating: float | None = p['rating']  # type: ignore
        ws[f'J{row}'] = '/' if rating is None else rating

        # 评论数
        review: int | None = p['review']  # type: ignore
        ws[f'K{row}'] = '/' if review is None else review

        # 最大可加购数
        max_qty: int = p['max_qty']  # type: ignore
        ws[f'L{row}'] = max_qty

    r = write_workbook_sync(save_path, wb)
    print(f'结果保存至 "{r}"')


if __name__ == '__main__':
    main()
