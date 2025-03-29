"""打包成 exe"""

from __future__ import annotations

import asyncio
from contextlib import asynccontextmanager
from math import ceil
from pathlib import Path
import subprocess
from typing import TYPE_CHECKING, overload

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from playwright.async_api import async_playwright
from scraper_utils.constants.workbook_style import (
    HYPERLINK_FONT,
    TEXT_CENTER_ALIGNMENT,
    RED_BOLD_FONT,
    YELLOW_FILL,
)
from scraper_utils.exceptions.browser_exception import PlaywrightError
from scraper_utils.utils.browser_util import abort_resources, ResourceType, MS1000
from scraper_utils.utils.json_util import write_json_sync, read_json_sync
from scraper_utils.utils.time_util import now_str
from scraper_utils.utils.workbook_util import write_workbook_sync, column_int2str as i2s

from emag_crawler.handlers.category_page import (
    goto_category_page,
    get_product_count_of_category,
    category_handler,
)
from emag_crawler.logger import logger as _logger
from emag_crawler.utils import build_category_page_url

if TYPE_CHECKING:
    from typing import AsyncGenerator, Literal

    from openpyxl.worksheet.worksheet import Worksheet
    from playwright.async_api import Browser, BrowserContext


cwd = Path.cwd()


def main():
    _logger.info('程序启动')

    # 输入要爬取的类目与其链接
    category, url = input_crawl_targe()

    # 启动 CDP
    launch_cdp()

    today = now_str('%m%d')
    json_save_dir = cwd / f'output/{category}/{today}'
    xlsx_save_path = cwd / f'output/{category}-{today}.xlsx'

    # 爬取数据
    asyncio.run(start_crawler(category, url, json_save_dir))

    # 将爬取的 json 数据保存成 xlsx
    wb, ws = create_workbook_template()
    products = read_product_json(json_save_dir)
    save_to_xlsx(wb, ws, products, xlsx_save_path)

    _logger.info('程序结束')


def launch_cdp(port: str = '9222') -> None:
    """启动 CDP"""

    chrome_path = Path('C:/Program Files/Google/Chrome/Application/chrome.exe')
    user_data_dir = (cwd / 'chrome_data').absolute()
    user_data_dir.mkdir(exist_ok=True)

    args = [
        str(chrome_path),
        f'--remote-debugging-port={port}',
        f'--user-data-dir={user_data_dir}',
        '--start-maximized',
        '--no-first-run',
        '--disable-sync',
        '--disable-default-apps',
        '--no-default-browser-check',
        'chrome://new-tab-page/',
    ]
    _logger.info(f'启动 CDP\n{' '.join(args)}')
    subprocess.Popen(
        args,
        shell=False,
        creationflags=subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP,
    )


async def start_crawler(category: str, first_page_url: str, json_save_dir: Path) -> None:
    async with connect_browser() as browser:
        context = browser.contexts[0]
        context.set_default_navigation_timeout(0)
        context.set_default_timeout(5 * MS1000)
        await abort_resources(
            context,
            (ResourceType.IMAGE, ResourceType.MEDIA, ResourceType.FONT),
        )

        # 爬取第 1 页
        product_count = await run_crawler(context, category, first_page_url, json_save_dir, 1)
        max_page_num = min(5, ceil(product_count / 60))
        _logger.debug(f'"{category}" 共有 {product_count} 个产品，最多爬取至第 {max_page_num} 页')

        # 爬取 2-[5] 页
        for i in range(2, max_page_num + 1):
            await run_crawler(context, category, first_page_url, json_save_dir, i)


def input_crawl_targe() -> tuple[str, str]:
    """输入要爬取的类目、类目页第一页的链接"""
    while True:
        category = input('要爬取的类目 >>> ')
        url = input('类目页第一页的链接 >>> ')

        if len(category) < 2 or len(url) < 2:
            print('输入的类目和类目页第一页的链接不对，重新输入')
            continue

        break

    return category, url


@asynccontextmanager
async def connect_browser(cdp_url: str = 'http://localhost:9222') -> AsyncGenerator[Browser]:
    """连接到 CDP"""
    async with async_playwright() as pwr:
        while True:
            try:
                browser = await pwr.chromium.connect_over_cdp(cdp_url, timeout=5 * MS1000)
            except PlaywrightError:
                _logger.error(f'无法连接至 CDP "{cdp_url}"')
                input('确认 CDP 启动后继续...')
                continue
            else:
                break

        try:
            yield browser
        finally:
            await browser.close()


@overload
async def run_crawler(
    context: BrowserContext,
    category: str,
    first_page_url: str,
    json_save_dir: Path,
    page_num: Literal[1] = 1,
) -> int: ...


@overload
async def run_crawler(
    context: BrowserContext, category: str, first_page_url: str, json_save_dir: Path, page_num: int
) -> None: ...


async def run_crawler(
    context: BrowserContext,
    category: str,
    first_page_url: str,
    json_save_dir: Path,
    page_num: int = 1,
):
    """
    爬取+保存爬取结果

    如果爬取的是第 1 页，会返回该类目有多少个产品
    """

    logger = _logger.bind(category=category)
    logger.info(f'爬取 "{category}" 的第 {page_num} 页')

    if page_num == 1:
        page = await goto_category_page(context, first_page_url, logger)
    else:
        page = await goto_category_page(context, build_category_page_url(first_page_url, page_num), logger)

    product_count = 0
    if page_num == 1:
        try:
            product_count = await get_product_count_of_category(page)
        except BaseException as be:
            logger.error(f'尝试解析 "{category}" 的产品总数时出错\n{be}')

    try:
        # 爬取数据
        result = await category_handler(page, category, logger)
    except BaseException as be:
        logger.error(f'爬取 "{category}" 的第 {page_num} 页时出错\n{be}')
    else:
        # 保存爬取结果为 json
        logger.info(f'保存 "{category}" 的第 {page_num} 页的爬取结果')
        json_save_dir.mkdir(parents=True, exist_ok=True)
        json_save_path = json_save_dir / f'{page_num}.json'
        save_path = write_json_sync(
            json_save_path,
            [_.model_dump() for _ in result],
            indent=4,
        )
        logger.success(f'"{category}" 的第 {page_num} 页的爬取结果已保存至 "{save_path}"')

    if page_num == 1:
        return product_count


def create_workbook_template():
    """创建表格模板"""
    wb = Workbook()
    ws: Worksheet = wb.active  # type: ignore

    # 标题行
    ws['A1'] = 'pnk'
    ws['B1'] = '详情页链接'
    ws['C1'] = '标题'
    ws['D1'] = '类目'
    ws['E1'] = '来源链接'
    ws['F1'] = '排名'
    ws['G1'] = '产品图'
    ws['H1'] = '价格'
    ws['I1'] = 'Top 标志'
    ws['J1'] = '评分'
    ws['K1'] = '评论数'
    ws['L1'] = '最大可加购数'

    # 设置标题行样式
    for c in range(1, 12 + 1):
        ws.cell(1, c).fill = YELLOW_FILL
        ws.cell(1, c).font = RED_BOLD_FONT
        ws.cell(1, c).alignment = TEXT_CENTER_ALIGNMENT
        ws.column_dimensions[i2s(c)].width = int(120 / 7)

    return wb, ws


def read_product_json(json_dir: Path):
    """读取产品的 json 数据"""
    result: list[dict[str, None | str | int | float | bool]] = list()

    files = json_dir.glob('*.json')

    for f in files:
        # 读取 json 文件
        data: list[dict[str, None | str | int | float | bool]] = read_json_sync(f)

        # 筛选加购失败的记录、保留需要的字段、将清理结果添加到 result
        for d in data:
            # 筛选加购失败的记录
            if d['cart_added'] is not True:
                continue

            # 保留需要的字段
            result.append(
                {
                    'pnk': d['pnk'],
                    'title': d['title'],
                    'category': d['category'],
                    'rank': d['rank_in_category'],
                    'source_url': d['source_url'],
                    'detail_url': d['detail_url'],
                    'image_url': d['image_url'],
                    'price': d['price'],
                    'top_favorite': d['top_favorite'],
                    'rating': d['rating'],
                    'review': d['review'],
                    'max_qty': d['max_qty'],
                }
            )

    # 按 rank 排序
    result.sort(key=lambda _: _['rank'])  # type: ignore

    return result.copy()


def save_to_xlsx(
    wb: Workbook,
    ws: Worksheet,
    data: list[dict[str, None | str | int | float | bool]],
    save_path: str | Path,
):
    """写入数据到表格，并保存表格为文件"""

    for row, p in enumerate(data, 2):
        # pnk
        pnk: str = p['pnk']  # type: ignore
        ws[f'A{row}'] = pnk

        # 详情页链接
        # detail_url = build_product_url(pnk)
        detail_url: str = p['detail_url']  # type: ignore
        ws[f'B{row}'] = detail_url
        ws[f'B{row}'].hyperlink = Hyperlink(ref=detail_url, target=detail_url)
        ws[f'B{row}'].font = HYPERLINK_FONT

        # 标题
        title: str = p['title']  # type: ignore
        ws[f'C{row}'] = title

        # 类目
        category: str = p['category']  # type: ignore
        ws[f'D{row}'] = category

        # 来源链接
        source_url: str = p['source_url']  # type: ignore
        ws[f'E{row}'] = source_url
        ws[f'E{row}'].hyperlink = Hyperlink(ref=source_url, target=source_url)
        ws[f'E{row}'].font = HYPERLINK_FONT

        # 排名
        rank: int = p['rank']  # type: ignore
        ws[f'F{row}'] = rank

        # 产品图
        image_url: str | None = p['image_url']  # type: ignore
        if image_url is not None:
            ws[f'G{row}'] = image_url
            ws[f'G{row}'].hyperlink = Hyperlink(ref=image_url, target=image_url)
            ws[f'G{row}'].font = HYPERLINK_FONT
        else:
            ws[f'G{row}'] = '/'

        # 价格
        price: float = p['price']  # type: ignore
        ws[f'H{row}'] = price

        # Top 标志
        top_favorite: bool = p['top_favorite']  # type: ignore
        ws[f'I{row}'] = '是' if top_favorite else '否'

        # 评分
        rating: float | None = p['rating']  # type: ignore
        ws[f'J{row}'] = '/' if rating is None else rating

        # 评论数
        review: int | None = p['review']  # type: ignore
        ws[f'K{row}'] = '/' if review is None else review

        # 最大可加购数
        max_qty: int = p['max_qty']  # type: ignore
        ws[f'L{row}'] = max_qty

    r = write_workbook_sync(save_path, wb)
    _logger.info(f'xlsx 保存至 "{r}"')


if __name__ == '__main__':
    main()
